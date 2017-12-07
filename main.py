import math
import os

from antlr4 import CommonTokenStream
from antlr4.FileStream import InputStream
from kivy.app import App
from kivy.core.window import Window
from kivy.graphics import Color, Rectangle, Line
from kivy.uix.button import Button
from kivy.uix.checkbox import CheckBox
from kivy.uix.dropdown import DropDown
from kivy.uix.filechooser import FileChooserListView
from kivy.uix.gridlayout import GridLayout
from kivy.uix.label import Label
from kivy.uix.popup import Popup
from kivy.uix.relativelayout import RelativeLayout
from kivy.uix.scrollview import ScrollView
from kivy.uix.textinput import TextInput
from openpyxl import Workbook
from openpyxl.styles import NamedStyle, Border, Side, PatternFill, Font

from mLexer import mLexer
from mParser import mParser
from mVisitor import mVisitor


class OutputTable(GridLayout):
    def __init__(self, **kwargs):
        super(OutputTable, self).__init__(**kwargs)

        self.background_width = None
        self.background_height = None

        # Colors for the output table
        self.base1 = list(map(lambda x: (x / 255), [0, 43, 54, 255]))
        self.base2 = list(map(lambda x: (x / 255), [7, 54, 66, 255]))
        self.error = list(map(lambda x: (x / 255), [220, 50, 47, 255]))
        self.pivot = list(map(lambda x: (x / 255), [133, 153, 0, 255]))
        self.highlight = list(map(lambda x: (x / 255), [38, 139, 210, 255]))

    def resize_background(self):
        self.canvas.before.clear()
        with self.canvas.before:
            Color(*self.base1)
            self.background = Rectangle(
                pos=(self.x, self.height * (self.number_of_iterations - self.iteration - 1)),
                size=self.size)
            Color(*self.base2)
            Rectangle(
                pos=(self.x, self.height * (self.number_of_iterations - self.iteration - 1) + self.height - 30),
                size=(self.width, 30))
            if self.entering_variable is not None:
                # Leaving variable's row
                Color(*self.highlight)
                self.leaving_variable_row_background = Rectangle(
                    pos=(200, self.height * (self.number_of_iterations - self.iteration - 1) + 30 * (
                        self.constraints - self.leaving_variable)),
                    size=(self.variables * 200, 30))
                # Entering variable's column
                Color(*self.highlight)
                self.entering_variable_row_background = Rectangle(pos=(
                    200 + self.entering_variable * 200,
                    30 + self.height * (self.number_of_iterations - self.iteration - 1)),
                    size=(200, self.constraints * 30))
                # Pivot
                Color(*self.pivot)
                self.pivot_background = Rectangle(
                    pos=(200 + 200 * self.entering_variable,
                         self.height * (self.number_of_iterations - self.iteration - 1) + 30 * (
                             self.constraints - self.leaving_variable)),
                    size=(200, 30))
            else:
                Color(*self.pivot)
                self.pivot_background = Rectangle(pos=(200 + 200 * self.variables, 30 + 30 * self.constraints),
                                                  size=(200, 30))


class SimplexOutputTable(OutputTable):
    def __init__(self, iteration, cj, variable_names, zj, cb, basic_vars, b, cj_minus_zj, entering_variable,
                 leaving_variable, a, number_of_iterations, **kwargs):
        super(SimplexOutputTable, self).__init__(**kwargs)
        self.cols = len(variable_names) + 3

        self.bind(minimum_height=self.setter('height'))
        self.bind(minimum_width=self.setter('width'))
        self.bind(minimum_height=lambda x, y: self.resize_background())
        self.bind(minimum_width=lambda x, y: self.resize_background())

        self.background = None
        self.pivot_background = None
        self.leaving_variable_row_background = None
        self.entering_variable_row_background = None
        self.entering_variable = entering_variable
        self.leaving_variable = leaving_variable
        self.variables = len(variable_names)
        self.constraints = len(cb)
        self.number_of_iterations = number_of_iterations
        self.iteration = iteration
        self.a = [x[:] for x in a]
        self.b = b[:]
        self.cb = cb[:]
        self.basic_vars = basic_vars[:]
        self.cj = cj[:]
        self.cj_minus_zj = cj_minus_zj[:]
        self.variable_names = variable_names[:]
        self.zj = zj[:]
        self.cj_minus_zj = cj_minus_zj[:]

        # First row: Iteration number and a padding of empty labels
        self.add_widget(Label(text=("Iter: " + str(iteration)), size_hint=(None, None), size=(100, 30)))
        for i in range(len(variable_names) + 2):
            self.add_widget(Label(size_hint=(None, None), size=(100, 30)))

        # Second row: 'Cj', and empty label, the actual values of Cj and another empty label
        self.add_widget(Label(text="Cj", size_hint=(None, None), size=(100, 30)))
        self.add_widget(Label(size_hint=(None, None), size=(100, 30)))
        for x in cj:
            self.add_widget(Label(text=x, size_hint=(None, None), size=(200, 30)))
        self.add_widget(Label(size_hint=(None, None), size=(100, 30)))

        # Third row: 'Cb', 'VarB' followed by the names of the variables and then 'bi'
        self.add_widget(Label(text="Cb", size_hint=(None, None), size=(100, 30)))
        self.add_widget(Label(text="VarB", size_hint=(None, None), size=(100, 30)))
        for x in variable_names:
            self.add_widget(Label(text=x, size_hint=(None, None), size=(200, 30)))
        self.add_widget(Label(text="bi", size_hint=(None, None), size=(200, 30)))

        # Fourth row: empty label, 'Zj', the actual values of Zj
        self.add_widget(Label(size_hint=(None, None), size=(100, 30)))
        self.add_widget(Label(text="Zj", size_hint=(None, None), size=(100, 30)))
        for i in range(len(zj)):
            self.add_widget(
                Label(text=(zj[i] if i < (len(zj) - 1) else zj[-1][:zj[-1].find("+")]), size_hint=(None, None),
                      size=(200, 30)))

        # Fifth to N-1th row: Cb[i], basic_variables[i], a[i][0]...a[i][len(a[i]) - 1], b[i]
        for i in range(len(a)):
            self.add_widget(Label(text=str(cb[i]), size_hint=(None, None), size=(100, 30)))
            self.add_widget(Label(text=str(basic_vars[i]), size_hint=(None, None), size=(100, 30)))
            for j in range(len(a[i])):
                new_label = Label(text=str(a[i][j]), size_hint=(None, None), size=(200, 30))
                self.add_widget(new_label)
            self.add_widget(
                Label(text=str(b[i]), size_hint=(None, None), size=(200, 30)))
        # Nth row: 'Cj - Zj', the actual value of cj_minus_zj and then an empty label
        self.add_widget(Label(text="Cj - Zj", size_hint=(None, None), size=(100, 30)))
        self.add_widget(Label(size_hint=(None, None), size=(100, 30)))
        for x in cj_minus_zj:
            self.add_widget(Label(text=x, size_hint=(None, None), size=(200, 30)))
        self.add_widget(Label(size_hint=(None, None), size=(100, 30)))


class AssignmentOutputTable(OutputTable):
    def __init__(self, costs, names, **kwargs):
        super(AssignmentOutputTable, self).__init__(**kwargs)
        self.n = len(costs)
        self.cols = self.n + 1

        self.bind(minimum_height=self.setter('height'))
        self.bind(minimum_width=self.setter('width'))
        self.bind(minimum_height=lambda x, y: self.resize_background())
        self.bind(minimum_width=lambda x, y: self.resize_background())

        self.add_widget(Label(text="Names", size_hint=(None, None), size=(200, 30)))
        for i in range(self.n):
            self.add_widget(Label(text=names[i], size_hint=(None, None), size=(200, 30)))

        for i in range(len(costs)):
            self.add_widget(Label(text=names[self.n + i], size_hint=(None, None), size=(200, 30)))
            for j in range(len(costs)):
                self.add_widget(Label(text=str(costs[i][j]), size_hint=(None, None), size=(200, 30)))

    def resize_background(self):
        self.canvas.before.clear()
        with self.canvas.before:
            Color(*self.base2)
            Rectangle(size=(200 * self.n, 30 * self.n),
                      pos=(200, self.y))


class AssignmentMinOutputTable(OutputTable):
    def __init__(self, costs, m, m_type, names, **kwargs):
        super(AssignmentMinOutputTable, self).__init__(**kwargs)
        self.n = len(costs)
        self.m_type = m_type
        self.cols = self.n + (2 if m_type == "Row" else 1)

        self.bind(minimum_height=self.setter('height'))
        self.bind(minimum_width=self.setter('width'))
        self.bind(minimum_height=lambda x, y: self.resize_background())
        self.bind(minimum_width=lambda x, y: self.resize_background())

        self.add_widget(Label(text="Names", size_hint=(None, None), size=(200, 30)))
        for i in range(self.n):
            self.add_widget(Label(text=names[i], size_hint=(None, None), size=(200, 30)))
        if m_type == "Row":
            self.add_widget(Label(text="Row. min", size_hint=(None, None), size=(200, 30)))

        for i in range(len(costs)):
            self.add_widget(Label(text=names[self.n + i], size_hint=(None, None), size=(200, 30)))
            for j in range(len(costs)):
                self.add_widget(Label(text=str(costs[i][j]), size_hint=(None, None), size=(200, 30)))
            if m_type == "Row":
                self.add_widget(Label(text=str(m[i]), size_hint=(None, None), size=(200, 30)))

        if m_type == "Column":
            self.add_widget(Label(text="Col. min", size_hint=(None, None), size=(200, 30)))
            for x in m:
                self.add_widget(Label(text=str(x), size_hint=(None, None), size=(200, 30)))

    def resize_background(self):
        self.canvas.before.clear()
        with self.canvas.before:
            Color(*self.base2)
            Rectangle(size=(200 * self.n, 30 * self.n),
                      pos=(200, self.y + (30 if self.m_type == "Column" else 0)))


class AssignmentLineOutputTable(OutputTable):
    def __init__(self, costs, names, h, v, **kwargs):
        super(AssignmentLineOutputTable, self).__init__(**kwargs)
        self.h = h
        self.v = v
        self.n = len(costs)
        self.cols = self.n + 1

        self.bind(minimum_height=self.setter('height'))
        self.bind(minimum_width=self.setter('width'))
        self.bind(minimum_height=lambda x, y: self.resize_background())
        self.bind(minimum_width=lambda x, y: self.resize_background())

        self.add_widget(Label(text="Names", size_hint=(None, None), size=(200, 30)))
        for i in range(self.n):
            self.add_widget(Label(text=names[i], size_hint=(None, None), size=(200, 30)))

        for i in range(len(costs)):
            self.add_widget(Label(text=names[self.n + i], size_hint=(None, None), size=(200, 30)))
            for j in range(len(costs)):
                self.add_widget(Label(text=str(costs[i][j]), size_hint=(None, None), size=(200, 30)))

    def resize_background(self):
        self.canvas.before.clear()
        self.canvas.after.clear()

        with self.canvas.before:
            Color(*self.base2)
            Rectangle(size=(200 * self.n, 30 * self.n),
                      pos=(200, self.y))

        for i in range(len(self.h)):
            if self.h[i][0]:
                with self.canvas.after:
                    Color(*self.error)
                    Line(points=(200, self.y + 30 * (len(self.h) - i - 1) + 14, 200 * (self.n + 1),
                                 self.y + 30 * (len(self.h) - i - 1) + 14), width=1)

        for i in range(len(self.v)):
            if self.v[i][0]:
                with self.canvas.after:
                    Color(*self.error)
                    Line(points=(200 * (i + 1) + 100, self.y + 30 * self.n, 200 * (i + 1) + 100, self.y), width=1)


class Constraint(GridLayout):
    def __init__(self, variables, **kwargs):
        super(Constraint, self).__init__(**kwargs)
        self.variables = []

        self.bind(minimum_height=self.setter('height'))
        self.bind(minimum_width=self.setter('width'))

        for x in range(1, variables + 1):
            self.variables.append(Variable(x, x == variables, cols=2,
                                           size_hint=(None, None), size=(200, 30)))

        for i in range(len(self.variables)):
            self.add_widget(self.variables[i])
        self.type = DropDown()
        for x in ["<=", ">=", "="]:
            btn = Button(text=x, size_hint_y=None, height=30,
                         border=(0, 0, 0, 0),
                         background_color=tuple(map(lambda x: (x / 255), [42, 161, 152, 192])),
                         background_normal="")
            btn.bind(on_release=lambda btn: self.type.select(btn.text))
            self.type.add_widget(btn)
        self.open_type = Button(text="<=", size_hint=(None, None), size=(50, 30),
                                border=(0, 0, 0, 0),
                                background_color=tuple(map(lambda x: (x / 255), [42, 161, 152, 192])),
                                background_normal="")
        self.open_type.bind(on_release=self.type.open)
        self.type.bind(on_select=lambda instance, x: setattr(self.open_type, "text", x))

        self.add_widget(self.open_type)

        self.rhs = TextInput(multiline=False,
                             size_hint=(None, None), size=(100, 30),
                             background_color=list(map(lambda x: (x / 255), [7, 54, 66, 255])),
                             border=(0, 0, 0, 0), foreground_color=(1, 1, 1, 1))
        self.add_widget(self.rhs)

    def get_coefficients(self):
        coefficients = []
        for x in self.variables:
            coefficients.append(x.get())
        return coefficients

    def get_rhs(self):
        try:
            return float(self.rhs.text)
        except ValueError:
            return 0.0

    def get_type(self):
        return self.open_type.text


class Variable(GridLayout):
    def __init__(self, i, last, **kwargs):
        super(Variable, self).__init__(**kwargs)

        self.bind(minimum_height=self.setter('height'))
        self.bind(minimum_width=self.setter('width'))
        self.var = TextInput(multiline=False, background_color=list(map(lambda x: (x / 255), [7, 54, 66, 255])),
                             border=(0, 0, 0, 0), foreground_color=(1, 1, 1, 1))
        self.add_widget(self.var)
        self.add_widget(Label(text=("X" + str(i) + ("" if last else " +"))))

    def get(self):
        try:
            return float(self.var.text)
        except ValueError:
            return 0.0


class SaveDialog(GridLayout):
    def __init__(self, save=None, dismiss=None):
        super(SaveDialog, self).__init__()
        self.cols = 1

        self.save = save
        self.dismiss = dismiss
        self.text_input = TextInput(size_hint_y=None, height=30, multiline=False,
                                    background_color=list(map(lambda x: (x / 255), [7, 54, 66, 255])),
                                    border=(0, 0, 0, 0), foreground_color=(1, 1, 1, 1))
        self.file_chooser = FileChooserListView()
        self.save_btn = Button(text="Save", size_hint_y=None, height=30,
                               background_color=tuple(map(lambda x: (x / 255), [42, 161, 152, 192])),
                               background_normal="",
                               border=(0, 0, 0, 0))
        self.save_btn.bind(on_release=lambda e: self.save(self.file_chooser.path, self.text_input.text))
        self.cancel_btn = Button(text="Cancel", size_hint_y=None, height=30,
                                 background_color=tuple(map(lambda x: (x / 255), [42, 161, 152, 192])),
                                 background_normal="",
                                 border=(0, 0, 0, 0))
        self.cancel_btn.bind(on_release=self.dismiss)

        self.add_widget(self.file_chooser)
        self.add_widget(self.text_input)
        self.add_widget(self.save_btn)
        self.add_widget(self.cancel_btn)


class Gui(GridLayout):
    def __init__(self, **kwargs):
        # create a default grid layout with custom width/height
        super(Gui, self).__init__(**kwargs)

        # when we add children to the grid layout, its size doesn't change at
        # all. we need to ensure that the height will be the minimum required
        # to contain all the childs. (otherwise, we'll child outside the
        # bounding box of the childs)
        self.bind(minimum_height=self.setter('height'))
        self.bind(minimum_width=self.setter('width'))

        self.base1 = list(map(lambda component: (component / 255), [0, 43, 54, 255]))
        self.base2 = list(map(lambda component: (component / 255), [7, 54, 66, 255]))
        self.base3 = list(map(lambda component: (component / 255), [253, 246, 227, 255]))
        self.error = list(map(lambda component: (component / 255), [220, 50, 47, 255]))
        self.pivot = list(map(lambda component: (component / 255), [133, 153, 0, 255]))
        self.highlight = list(map(lambda component: (component / 255), [38, 139, 210, 255]))

        # self.bind(minimum_height=lambda x, y: self.resize_background())
        # self.bind(minimum_width=lambda x, y: self.resize_background())

        # How many variables has the problem?
        self.add_widget(
            Label(text="For LP and ILP problems", size_hint=(None, None), size=(400, 30)))
        self.add_widget(
            Label(text="", size_hint=(None, None), size=(400, 30)))

        self.add_widget(
            Label(text="How many decision variables has the problem?", size_hint=(None, None), size=(400, 30)))
        self.number_of_variables = TextInput(size_hint=(None, None), size=(400, 30), background_color=self.base2,
                                             border=(0, 0, 0, 0), foreground_color=(1, 1, 1, 1))
        self.add_widget(self.number_of_variables)

        # How many constraints?
        self.add_widget(
            Label(text="How many constraints?"))
        self.number_of_constraints = TextInput(size_hint=(None, None), size=(400, 30), background_color=self.base2,
                                               border=(0, 0, 0, 0), foreground_color=(1, 1, 1, 1))
        self.add_widget(self.number_of_constraints)

        self.add_widget(
            Label(text="", size_hint=(None, None), size=(400, 30)))

        self.add_widget(
            Label(text="", size_hint=(None, None), size=(400, 30)))

        # Assignment problems input
        self.add_widget(
            Label(text="For assignment problems", size_hint=(None, None), size=(400, 30)))
        self.add_widget(
            Label(text="", size_hint=(None, None), size=(400, 30)))

        self.add_widget(
            Label(text="Size of the cost matrix"))
        self.cost_matrix_size = TextInput(size_hint=(None, None), size=(400, 30), background_color=self.base2,
                                          border=(0, 0, 0, 0), foreground_color=(1, 1, 1, 1))
        self.add_widget(self.cost_matrix_size)

        # Type of the problem
        self.problem_nature = DropDown()
        for x in ["Linear Programming", "Integer Linear Programming", "Assignment Problem", "Transportation Problem"]:
            btn = Button(text=x, size_hint_y=None, height=30,
                         background_color=tuple(map(lambda x: (x / 255), [42, 161, 152, 192])),
                         background_normal="",
                         border=(0, 0, 0, 0))
            btn.bind(on_release=lambda btn: self.problem_nature.select(btn.text))
            self.problem_nature.add_widget(btn)
        self.open_problem_nature = Button(text="Linear Programming", size_hint=(None, None), size=(400, 30),
                                          background_color=tuple(map(lambda x: (x / 255), [42, 161, 152, 192])),
                                          background_normal="",
                                          border=(0, 0, 0, 0))
        self.open_problem_nature.bind(on_release=self.problem_nature.open)
        self.problem_nature.bind(on_select=lambda instance, x: setattr(self.open_problem_nature, "text", x))

        self.add_widget(Label(text="Type of the problem:", size_hint=(None, None), size=(400, 30)))
        self.add_widget(self.open_problem_nature)

        self.submit = Button(text="Continue", size_hint=(None, None), size=(400, 30),
                             background_color=tuple(map(lambda x: (x / 255), [42, 161, 152, 192])),
                             background_normal="",
                             border=(0, 0, 0, 0))
        self.submit.bind(on_release=self.gen_input_table)
        self.add_widget(self.submit)

        self.fo = None
        self.type = None
        self.open_problem_type = None
        self.integer_check = []
        self.constraints = []
        self.constraints_types = []
        self._popup = None

    def resize_background(self):
        self.canvas.before.clear()
        with self.canvas.before:
            Color(*self.base1)
            Rectangle(
                pos=self.pos,
                size=self.size)

    def gen_input_table(self, e):
        self.clear_widgets()
        if self.open_problem_nature.text == "Integer Linear Programming" or \
                        self.open_problem_nature.text == "Linear Programming":
            self.cols = 1
            if self.number_of_variables.text == "":
                self.solve(None)
                return
            variables = int(self.number_of_variables.text)
            constraints = int(self.number_of_constraints.text)
            self.type = DropDown()
            for x in ["Maximize", "Minimize"]:
                btn = Button(text=x, size_hint_y=None, height=30,
                             background_color=tuple(map(lambda x: (x / 255), [42, 161, 152, 192])),
                             background_normal="",
                             border=(0, 0, 0, 0))
                btn.bind(on_release=lambda btn: self.type.select(btn.text))
                self.type.add_widget(btn)
            self.open_problem_type = Button(text="Maximize", size_hint=(None, None), size=(100, 30),
                                            background_color=tuple(map(lambda x: (x / 255), [42, 161, 152, 192])),
                                            background_normal="",
                                            border=(0, 0, 0, 0))
            self.open_problem_type.bind(on_release=self.type.open)
            self.type.bind(on_select=lambda instance, x: setattr(self.open_problem_type, "text", x))

            self.add_widget(self.open_problem_type)

            self.fo = GridLayout(cols=(variables + 1), padding=10, spacing=10,
                                 size_hint=(None, None), width=Window.width)

            self.fo.bind(minimum_height=self.setter('height'))
            self.fo.bind(minimum_width=self.setter('width'))

            self.fo.add_widget(Label(text="Objective funtion: ", size_hint=(None, None), size=(150, 30)))
            for i in range(1, variables + 1):
                self.fo.add_widget(Variable(i, i == variables, cols=2,
                                            size_hint=(None, None), size=(200, 30)))

            self.add_widget(self.fo)

            self.add_widget(Label(text="Subject to:", size_hint=(None, None), size=(200, 50)))

            self.constraints = []

            for i in range(1, constraints + 1):
                self.constraints.append(Constraint(variables, cols=(variables + 2), padding=10, spacing=10,
                                                   size_hint=(None, None), width=Window.width))
                self.add_widget(self.constraints[i - 1])

            self.integer_check = []
            if self.open_problem_nature.text == "Integer Linear Programming":
                self.add_widget(
                    Label(text="The following variables must be integers:", size_hint=(None, None), size=(400, 30)))
                for i in range(1, variables + 1):
                    new_container = GridLayout(cols=2, size_hint=(None, None), width=125)
                    new_container.add_widget(Label(text=("x" + str(i)), size=(100, 30)))
                    new_check_box = CheckBox()
                    self.integer_check.append(new_check_box)
                    new_container.add_widget(new_check_box)
                    self.add_widget(new_container)

            self.submit = Button(size_hint=(None, None), size=(400, 30),
                                 background_color=tuple(map(lambda x: (x / 255), [42, 161, 152, 192])),
                                 background_normal="",
                                 border=(0, 0, 0, 0))
            self.submit.bind(on_release=self.solve)
            self.submit.text = "Continue"
            self.add_widget(self.submit)

        elif self.open_problem_nature.text == "Assignment Problem":
            if self.cost_matrix_size.text == "":
                self.solve(None)
                return
            n = int(self.cost_matrix_size.text)
            self.cols = n + 1
            self.add_widget(Label(text="Names", size_hint=(None, None), size=(200, 30)))
            for i in range(n):
                self.add_widget(TextInput(size_hint=(None, None), size=(200, 30), multiline=False,
                                          background_color=list(map(lambda x: (x / 255), [7, 54, 66, 255])),
                                          border=(0, 0, 0, 0), foreground_color=(1, 1, 1, 1)))
            for i in range((int(n) + 1) * n):
                self.add_widget(TextInput(size_hint=(None, None), size=(200, 30), multiline=False,
                                          background_color=list(map(lambda x: (x / 255), [7, 54, 66, 255])),
                                          border=(0, 0, 0, 0), foreground_color=(1, 1, 1, 1)))
            self.add_widget(Button(text="Continue", on_release=self.solve, size_hint=(None, None), size=(200, 30),
                                   background_color=tuple(map(lambda x: (x / 255), [42, 161, 152, 192])),
                                   background_normal="",
                                   border=(0, 0, 0, 0)))

        elif self.open_problem_nature.text == "Transportation Problem":
            self.cols = 1
            self.solve(None)

    def solve(self, e):
        problem_nature = self.open_problem_nature.text
        if problem_nature == "Linear Programming":
            problem_type = self.open_problem_type.text
            variables = int(self.number_of_variables.text)
            constraints = int(self.number_of_constraints.text)
            self.clear_widgets()
            # Coefficients of the objective function
            cj = [str(x.get()) for x in
                  filter(lambda y: type(y) == Variable, self.fo.children)]
            cj.reverse()
            a = [x.get_coefficients() for x in self.constraints]  # Coefficients of the constraints' variables
            b = [x.get_rhs() for x in self.constraints]  # Rhs
            constraint_types = [x.get_type() for x in self.constraints]

            # We need to "fix" every constraint with a negative bi
            for i in range(len(b)):
                if b[i] < 0.0:
                    a[i] = [-x for x in a[i]]
                    b[i] = -b[i]
                    if constraint_types[i] == ">=":
                        constraint_types[i] = "<="
                    elif constraint_types[i] == "<=":
                        constraint_types[i] = ">="

            variable_names = ["x" + str(i) for i in range(1, variables + 1)]
            iteration = 0

            s = Simplex(variables, constraints, cj, a, b, constraint_types, variable_names, problem_type)
            output_tables, solution, basic_vars = s.solve()
            self.add_widget(Label(text="Final solution:", size_hint=(None, None), size=(100, 30)))
            self.add_widget(
                Label(text=(("Zmax = " if problem_type == "Maximize" else "Zmin = ") + str(solution[0])),
                      size_hint=(None, None),
                      size=(250, 30)))
            basic_vars = [(x[:-1] + "x" if x[-1] == "h" else x) for x in basic_vars]
            for i in range(1, len(solution)):
                self.add_widget(Label(text=("x" + str(i) + " = " + str(solution[i])), size_hint=(None, None),
                                      size=(250, 30)))

            export_buttons = GridLayout(size=(250, 60), cols=1, size_hint=(None, None))

            save_to_xl = Button(text="Export as Excel 2010 xlsx", size_hint=(None, None), size=(250, 30),
                                background_color=tuple(map(lambda x: (x / 255), [42, 161, 152, 192])),
                                background_normal="",
                                border=(0, 0, 0, 0))
            save_to_xl.bind(on_release=lambda e: self.show_save_xl(e, s))
            export_buttons.add_widget(save_to_xl)
            save_to_png = Button(text="Export as png", size_hint=(None, None), size=(250, 30),
                                 background_color=tuple(map(lambda x: (x / 255), [42, 161, 152, 192])),
                                 background_normal="",
                                 border=(0, 0, 0, 0))
            save_to_png.bind(on_release=self.show_save_png)
            export_buttons.add_widget(save_to_png)

            self.add_widget(export_buttons)

            for i in range(len(output_tables)):
                self.add_widget(output_tables[i])
        elif problem_nature == "Integer Linear Programming":  # Integer linear
            if self.number_of_variables.text == "":
                problem_type = "Maximize"
                variables = 4
                constraints = 4
                self.clear_widgets()
                # Coefficients of the objective function
                cj = ["4.0", "-2.0", "7.0", "-1.0"]
                a = [[1.0, 0.0, 5.0, 0.0], [1.0, 1.0, -1.0, 0.0], [6.0, -5.0, 0.0, 0.0], [-1.0, 0.0, 2.0, -2.0]]
                b = [10.0, 1.0, 0.0, 3.0]
                constraint_types = ["<=", "<=", "<=", "<="]

                variable_names = ["x" + str(i) for i in range(1, variables + 1)]
                iteration = 0

                integer_check = [1, 1, 1, 0]
                height = 0
                # Initial problem
                problems = [[Simplex(variables, constraints, cj[:], [x[:] for x in a], b[:], constraint_types[:],
                                     variable_names, problem_type)]]
            else:
                problem_type = self.open_problem_type.text
                variables = int(self.number_of_variables.text)
                constraints = int(self.number_of_constraints.text)
                self.clear_widgets()
                # Coefficients of the objective function
                cj = [str(x.get()) for x in
                      filter(lambda y: type(y) == Variable, self.fo.children)]
                cj.reverse()
                a = [x.get_coefficients() for x in self.constraints]  # Coefficients of the constraints' variables
                b = [x.get_rhs() for x in self.constraints]  # Rhs
                constraint_types = [x.get_type() for x in self.constraints]

                variable_names = ["x" + str(i) for i in range(1, variables + 1)]
                iteration = 0

                integer_check = [x.active for x in self.integer_check]
                height = 0
                # Initial problem
                problems = [[Simplex(variables, constraints, cj[:], [x[:] for x in a], b[:], constraint_types[:],
                                     variable_names, problem_type)]]

            # We need to "fix" every constraint with a negative bi
            for i in range(len(b)):
                if b[i] < 0.0:
                    a[i] = [-x for x in a[i]]
                    b[i] = -b[i]
                    if constraint_types[i] == ">=":
                        constraint_types[i] = "<="
                    elif constraint_types[i] == "<=":
                        constraint_types[i] = ">="

            optimal_solution_aux = -math.inf if problem_type == "Maximize" else math.inf
            optimal_solution = None
            optimal_solution_coord = []
            while True:
                for x in problems[height]:
                    if x is not None:
                        x.solve()

                problems.append([])
                for i in range(len(problems[height])):
                    if problems[height][i] is None:
                        problems[height + 1].append(None)
                        problems[height + 1].append(None)
                    else:
                        solution = problems[height][i].solution
                        if solution is None:
                            problems[height + 1].append(None)
                            problems[height + 1].append(None)
                        else:
                            new_iter = False
                            for k in range(1, len(solution)):
                                if integer_check[k - 1] and not solution[k].is_integer():
                                    new_iter = True
                                    np1 = Simplex.from_simplex(problems[height][i]).add_constraint(
                                        [0.0 for l in range(k - 1)] + [1.0] + [0.0 for l in
                                                                               range(len(solution) - k - 1)],
                                        "<=",
                                        int(solution[k]))
                                    np2 = Simplex.from_simplex(problems[height][i]).add_constraint(
                                        [0.0 for l in range(k - 1)] + [1.0] + [0.0 for l in
                                                                               range(len(solution) - k - 1)],
                                        ">=",
                                        int(solution[k]) + 1)
                                    problems[height + 1].append(np1)
                                    problems[height + 1].append(np2)
                                    break
                            if not new_iter:
                                if problem_type == "Maximize":
                                    if optimal_solution_aux < solution[0]:
                                        optimal_solution_aux = solution[0]
                                        optimal_solution = solution
                                        optimal_solution_coord = [height, i]
                                else:
                                    if optimal_solution_aux > solution[0]:
                                        optimal_solution_aux = solution[0]
                                        optimal_solution = solution
                                        optimal_solution_coord = [height, i]
                                problems[height + 1].append(None)
                                problems[height + 1].append(None)
                if not any(problems[height + 1]):
                    break
                height += 1

            # It's time to print the tree to the gui
            text_height = 30 * (variables + 1)
            text_width = 200
            sibling_spacing = text_width
            layout = RelativeLayout(size_hint=(None, None), size=(
                text_width * len(problems[-1]) + text_width, text_height * len(problems) + 100 * len(problems)))
            with self.canvas.before:
                Color(*self.base1)
                Rectangle(pos=layout.pos, size=layout.size)

            def toggle(x, p):
                nonlocal layout
                if not x.status:
                    x.status = 1
                    x.text = "-"
                    if x.start:
                        x.start = False
                        x.scroll_view = ScrollView(size_hint=(None, None), size=(500, 300),
                                                   do_scroll_x=True, pos=(
                                x.x + x.width * 2, x.y - 300 - x.height))
                        gl = GridLayout(cols=1,
                                        size_hint=(None, None), width=500)
                        gl.bind(minimum_width=gl.setter('width'))
                        gl.bind(minimum_height=gl.setter('height'))

                        for table in p[x.i][x.j].output_tables:
                            gl.add_widget(table)
                        x.scroll_view.add_widget(gl)
                        layout.add_widget(x.scroll_view)
                    else:
                        layout.add_widget(x.scroll_view)
                else:
                    x.status = 0
                    x.text = "+"
                    layout.remove_widget(x.scroll_view)

            for i in range(len(problems) - 1):
                for j in range(len(problems[i])):
                    if problems[i][j] is not None:
                        base_x = ((sibling_spacing * j * (len(problems[-1])) /
                                   len(problems[i])) + sibling_spacing * (
                                      len(problems[-1]) / (2 ** (i + 1))))

                        base_y = text_height * (len(problems) - i - 1) + 100 * (len(problems) - i - 1)

                        if i != len(problems) - 2 and (
                                        problems[i + 1][j * 2] is None or problems[i][j].solution is None) and i != 0:
                            base_x = ((sibling_spacing * (j // 2) * (len(problems[-1])) /
                                       len(problems[i - 1])) + sibling_spacing * (
                                          len(problems[-1]) / (2 ** i))) + (
                                         text_width * 2 if (j + 1) % 2 == 0 else - text_width * 2)

                        if problems[i][j].solution is None:
                            text = "infeasible"
                            color = self.error
                        else:
                            color = self.base2
                            if i == optimal_solution_coord[0] and j == optimal_solution_coord[1]:
                                color = self.highlight
                            text = ("Zmax = " if problem_type == "Maximize" else "Zmin = ") + str(
                                problems[i][j].solution[
                                    0]) + "\n"
                            for k in range(1, len(problems[i][j].solution)):
                                text += "x" + str(k) + " = " + str(problems[i][j].solution[k]) + "\n"

                        if i != 0:
                            layout.add_widget(
                                Label(text=(
                                    "x" + str(problems[i][j].a[-1].index(1.0) + 1) + " " +
                                    problems[i][j].constraint_types[
                                        -1] + " " + str(problems[i][j].b[-1])),
                                    pos=(
                                        base_x,
                                        base_y + text_height / 2 + 15),
                                    size=(text_width, text_height), size_hint=(None, None)))

                            with self.canvas.before:
                                Color(*self.base3)
                                Line(points=(
                                    base_x + text_width / 2,
                                    base_y + text_height,
                                    text_width / 2 +
                                    ((sibling_spacing * (j // 2) * (len(problems[-1])) /
                                      len(problems[i - 1])) + sibling_spacing * (
                                         len(problems[-1]) / (2 ** i))), text_height +
                                    base_y + 100), width=1)

                        with self.canvas.before:
                            Color(*color)
                            Rectangle(pos=(base_x,
                                           base_y),
                                      size=(text_width, text_height))

                            Color(*self.base3)
                            Line(points=(
                                base_x, base_y, base_x, base_y + text_height,
                                base_x + text_width,
                                base_y + text_height, base_x + text_width, base_y, base_x,
                                base_y, base_x + text_width, base_y), width=1)

                        new_button = Button(text="+", size_hint=(None, None),
                                            size=(30, 30),
                                            pos=(base_x + text_width - 15, base_y + text_height - 17),
                                            border=(0, 0, 0, 0),
                                            background_color=tuple(map(lambda x: (x / 255), [42, 161, 152, 192])))
                        new_button.i = i
                        new_button.j = j
                        new_button.background_normal = ""
                        new_button.background_down = ""
                        new_button.start = True
                        new_button.status = 0
                        new_button.bind(on_release=lambda x: toggle(x, problems))
                        layout.add_widget(new_button)

                        layout.add_widget(Label(text=text,
                                                pos=(
                                                    base_x,
                                                    base_y),
                                                size=(text_width, text_height), size_hint=(None, None)))
            # Printing the optimal solution
            text = "Optimal solution\n" + ("Zmax = " if problem_type == "Maximize" else "Zmin = ") + str(
                optimal_solution[0]) + "\n"
            for i in range(1, len(optimal_solution)):
                text += "x" + str(i) + " = " + str(optimal_solution[i]) + "\n"
            new_label = Label(text=text, size=(text_width, text_height),
                              pos=(0, len(problems) * text_height + 100 * (len(problems) - 2)), size_hint=(None, None))
            with self.canvas.before:
                Color(*self.highlight)
                Rectangle(pos=new_label.pos, size=new_label.size)
            layout.add_widget(new_label)

            save_btn = Button(text="Export as png", size=(text_width, 30),
                              pos=(0, len(problems) * text_height - 30 + 100 * (len(problems) - 2)),
                              size_hint=(None, None),
                              border=(0, 0, 0, 0),
                              background_color=tuple(map(lambda x: (x / 255), [42, 161, 152, 192])),
                              background_normal=""
                              )
            save_btn.bind(on_release=self.show_save_png)
            layout.add_widget(save_btn)
            self.add_widget(layout)
        elif problem_nature == "Assignment Problem":
            self.cols = 1

            def isfloat(value):
                try:
                    float(value)
                    return True
                except ValueError:
                    return False

            if self.cost_matrix_size.text != "":
                n = int(self.cost_matrix_size.text)
                names = [x.text for x in
                         filter(lambda y: type(y) == TextInput and not isfloat(y.text), self.children)]
                names.reverse()
                values = [float(x.text) for x in
                          filter(lambda y: type(y) == TextInput and isfloat(y.text), self.children)]
                values.reverse()
                vaux = values
                values = [[vaux[j] for j in range(i, i + n)] for i in range(0, len(vaux), n)]

            else:
                n = 3
                names = ["x", "y", "z", "a", "b", "c"]
                values = [[10.0, 9.0, 5.0], [9.0, 8.0, 3.0], [6.0, 4.0, 7.0]]

            self.clear_widgets()
            # Iterate until the job is done

            output = []
            height = 30 * (n + 3)

            # 1. Row min
            rm = [min(x) for x in values]

            output.append(
                Label(text="1. Subtract Row Minimal", size_hint=(None, None), size=(200, 30)))

            output.append(
                AssignmentMinOutputTable(values, rm, "Row", names, size_hint=(None, None), width=Window.width))

            for i in range(len(values)):
                if rm[i] != 0:
                    for j in range(len(values)):
                        values[i][j] -= rm[i]

            # 2. Col min
            cm = [min(x) for x in [[values[j][i] for j in range(n)] for i in range(n)]]

            output.append(Label(text="2. Subtract Column Minimal", size_hint=(None, None), size=(200, 30)))

            output.append(
                AssignmentMinOutputTable(values, cm, "Column", names, size_hint=(None, None), width=Window.width))

            for i in range(len(values)):
                if cm[i] != 0:
                    for j in range(len(values)):
                        values[j][i] -= cm[i]

            while True:
                # 3. Lines
                h = [[1 if 0.0 in x else 0, x.count(0.0)] for x in values]  # horizontal lines
                # Vertical lines
                v = [[1 if 0.0 in x else 0, x.count(0.0)] for x in
                     [[values[j][i] for j in range(n)] for i in range(n)]]
                # Removing unnecessary lines
                for i in range(len(v)):
                    if v[i][1] == 1:
                        for j in range(len(h)):
                            if values[j][i] == 0 and (h[j][1] > 1):
                                v[i][0] = 0
                                break
                for i in range(len(h)):
                    for j in range(len(v)):
                        if values[i][j] == 0 and v[j][0] == 1:
                            h[i][1] -= 1
                    if h[i][1] == 0:
                        h[i][0] = 0

                output.append(Label(text="3. Covering the zeroes with lines",
                                    size_hint=(None, None), size=(220, 30)))
                output.append(
                    AssignmentLineOutputTable(values, names, h, v, size_hint=(None, None), width=Window.width))

                # If the number of lines is equal to n, we stop, if not, we continue with step 4
                lines = 0
                for i in range(len(values)):
                    lines += v[i][0] + h[i][0]
                if lines == len(values):
                    break

                output.append(Label(text="4. Creating additional zeroes",
                                    size_hint=(None, None), size=(200, 30)))
                # 4. Create additional zeros
                # 4.1 minimum value not covered by a line
                min_value = math.inf
                for i in range(len(values)):
                    for j in range(len(values)):
                        if (v[j][0] + h[i][0]) == 0 and min_value > values[i][j]:
                            min_value = values[i][j]
                # 4.2 subtract the minimum value from all uncovered values (and add it to the elements
                # covered by two lines)
                for i in range(len(values)):
                    for j in range(len(values)):
                        if (v[j][0] + h[i][0]) == 0:
                            values[i][j] -= min_value
                        elif (v[j][0] + h[i][0]) == 2:
                            values[i][j] += min_value

                output.append(
                    AssignmentOutputTable(values, names, size_hint=(None, None), width=Window.width))

            for i in range(len(output)):
                if type(output[i]) == Label:
                    output[i].pos = (0, (len(output) - i) * height + 30)
                else:
                    output[i].pos = (0, (len(output) - i) * height - 30)

            layout = RelativeLayout(size=(400 + 200 * n, height * len(output) + 30 * (n + 1)), size_hint=(None, None))
            for x in output:
                layout.add_widget(x)

            solution = []
            zeroes = [x.count(0.0) for x in values]
            min_number_of_zeroes = math.inf

            def min_z():
                nonlocal min_number_of_zeroes
                nonlocal values
                for i in range(len(values)):
                    z = values[i].count(0.0)
                    if i not in solution and z < min_number_of_zeroes and z != 0:
                        min_number_of_zeroes = z

            while len(solution) < len(values):

                for i in range(len(values)):
                    cmp = [x[1] for x in solution]

                    for j in range(len(values)):
                        if values[i][j] == 0.0 and j not in cmp and zeroes[
                            i] == min_number_of_zeroes:
                            solution.append([i, j])
                            cmp.append(j)
                            values[i][j] = -1
                            for k in range(j + 1, len(values)):
                                if values[i][k] == 0.0:
                                    values[i][k] = -1
                            zeroes[i] = math.inf
                            break

                    for j in range(len(values)):
                        for k in range(len(values)):
                            if values[j][k] == 0.0 and k in cmp:
                                zeroes[j] -= 1
                                values[j][k] = -1

                    min_number_of_zeroes = math.inf
                    min_z()

            solution.sort(key=lambda x: x[0])

            layout.add_widget(Label(text="Solution:", size=(200, 30), size_hint=(None, None), pos=(0, 30 * (n + 1))))
            for i in range(len(solution)):
                layout.add_widget(
                    Label(text=names[len(values) + solution[i][0]], size=(200, 30), pos=(0, 30 * (n - i)),
                          size_hint=(None, None)))
                layout.add_widget(
                    Label(text=names[solution[i][1]], size=(200, 30), pos=(400, 30 * (n - i)), size_hint=(None, None)))

            self.canvas.before.clear()
            with self.canvas.before:
                Color(*self.base1)
                Rectangle(
                    pos=layout.pos,
                    size=layout.size)

            for x in solution:
                with layout.canvas.before:
                    Color(*self.error)
                    Line(points=(200, 30 * (n - x[0]) + 14, 400, 30 * (n - x[0]) + 14, 400, 30 * (n - x[0]) + 21, 407,
                                 30 * (n - x[0]) + 14, 400, 30 * (n - x[0]) + 7, 400, 30 * (n - x[0]) + 14), width=1)

            save_to_png = Button(text="Export as png", size_hint=(None, None), size=(200, 30),
                                 background_color=tuple(map(lambda x: (x / 255), [42, 161, 152, 192])),
                                 background_normal="",
                                 border=(0, 0, 0, 0),
                                 pos=(200, 0))
            save_to_png.bind(on_release=self.show_save_png)
            layout.add_widget(save_to_png)
            self.add_widget(layout)
        elif problem_nature == "Transportation Problem":
            costs = [[10, 0, 20, 11], [12, 7, 9, 20], [0, 14, 16, 18]]
            assignment = [[0 for i in range(len(x))] for x in costs]
            demand = [5, 15, 15, 10]
            supply = [15, 25, 5]
            origin_names = ["1", "2", "3"]
            destination_names = ["1", "2", "3", "4"]

            e_rows = []
            e_cols = []
            while True:
                # 1 Calculating the penalizations per column and row
                p_row = []
                p_col = []
                for i in range(len(costs)):
                    if i not in e_rows:
                        min_cost = min(costs[i])
                        second_min_cost = min(filter(lambda x: x != min_cost, costs[i]))
                        p_row.append(second_min_cost - min_cost)
                    else:
                        p_row.append(-math.inf)

                for i in range(len(costs[0])):
                    if i not in e_cols:
                        min_cost = math.inf
                        second_min_cost = math.inf
                        for j in range(len(costs)):
                            if min_cost > costs[j][i]:
                                min_cost = costs[j][i]

                        for j in range(len(costs)):
                            if second_min_cost > costs[j][i] != min_cost:
                                second_min_cost = costs[j][i]

                        p_col.append(second_min_cost - min_cost)
                    else:
                        p_col.append(-math.inf)

                print("1. Penalization")
                print("Costs: ", costs)
                print("Demand: ", demand)
                print("Supply: ", supply)
                print("row pen: ", p_row)
                print("col pen: ", p_col)
                pr = max(p_row)
                pc = max(p_col)
                print("max row pen: ", pr)
                print("max col pen: ", pc)
                if len(e_rows) == len(e_cols) == len(costs) or pr == pc == -math.inf:
                    break

                if pr >= pc:
                    row = p_row.index(pr)
                    c = []
                    for i in range(len(costs[row])):
                        if i in e_cols:
                            c.append(math.inf)
                        else:
                            c.append(costs[row][i])
                    column = c.index(min(c))

                    difference = supply[row] - demand[column]
                    if difference == 0:
                        assignment[row][column] = demand[column]
                        supply[row] = demand[column] = 0
                    elif difference < 0:
                        assignment[row][column] = supply[row]
                        demand[column] = 0
                        supply[row] = abs(difference)
                    else:
                        assignment[row][column] = demand[column]
                        supply[row] = difference
                        demand[column] = 0

                    if demand[column] == 0:
                        e_cols.append(column)
                    if supply[row] == 0:
                        e_rows.append(row)
                else:
                    column = p_col.index(pc)
                    c = []
                    for i in range(len(costs)):
                        if i in e_rows:
                            c.append(math.inf)
                        else:
                            c.append(costs[i][column])
                    row = c.index(min(c))
                    difference = supply[row] - demand[column]

                    if difference == 0:
                        assignment[row][column] = demand[column]
                        supply[row] = demand[column] = 0
                    elif difference < 0:
                        assignment[row][column] = supply[row]
                        demand[column] = 0
                        supply[row] = abs(difference)
                    else:
                        assignment[row][column] = demand[column]
                        supply[row] = difference
                        demand[column] = 0

                    if demand[column] == 0:
                        e_cols.append(column)
                    if supply[row] == 0:
                        e_rows.append(row)
                print("Assignment: ", assignment)
                print("finished rows: ", e_rows)
                print("finished cols: ", e_cols)

            print("STARTING M.O.D.I")
            while True:
                # 1. getting the values for the ui and vj
                u = [None for x in costs]
                v = [None for x in costs[0]]
                while True:
                    pass

    def show_save_png(self, btn):
        content = SaveDialog(save=lambda path, filename: self.save_png(btn, path, filename), dismiss=self.dismiss_popup)
        self._popup = Popup(title="Save file", content=content,
                            size_hint=(0.9, 0.9),
                            background="src/images/colors/base1.png")
        self._popup.open()

    def dismiss_popup(self, e):
        self._popup.dismiss()

    def save_png(self, btn, path, filename):
        draw_inst = None

        if type(btn.parent) == RelativeLayout:
            draw_inst = btn.canvas.children[:]
            btn.canvas.clear()
        else:
            draw_inst = btn.parent.canvas.children[:]
            btn.parent.canvas.clear()
        f = os.path.join(path, filename)
        self.export_to_png(f + ".png" if filename[-5:] != ".png" else f)
        self._popup.dismiss()

        if type(btn.parent) == RelativeLayout:
            for x in draw_inst:
                btn.parent.canvas.add(x)
        else:
            self.resize_background()
            for x in draw_inst:
                btn.parent.canvas.add(x)

    def show_save_xl(self, btn, problem):
        content = SaveDialog(save=lambda path, filename: self.save_xl(path, filename, problem),
                             dismiss=self.dismiss_popup)
        self._popup = Popup(title="Save file", content=content,
                            size_hint=(0.9, 0.9),
                            background="src/images/colors/base1.png")
        self._popup.open()

    def save_xl(self, path, filename, problem):
        f = os.path.join(path, filename) + (".xlsx" if filename[-5:] != ".xslx" else "")
        wb = Workbook()
        ws = wb.active
        self._popup.dismiss()

        base = NamedStyle(name="base")
        bd = Side(style='thin', color="586e75")
        base.border = Border(left=bd, top=bd, right=bd, bottom=bd)
        base.fill = PatternFill(start_color='00002b36',
                                end_color='00002b36',
                                fill_type='solid')
        base.font = Font(color="fdf6e3")
        wb.add_named_style(base)

        base2 = NamedStyle(name="base2")
        bd = Side(style='thin', color="586e75")
        base2.border = Border(left=bd, top=bd, right=bd, bottom=bd)
        base2.fill = PatternFill(start_color='00073642',
                                 end_color='00073642',
                                 fill_type='solid')
        base2.font = Font(color="fdf6e3")
        wb.add_named_style(base2)

        highlight = NamedStyle(name="highlight")
        bd = Side(style='thin', color="586e75")
        highlight.border = Border(left=bd, top=bd, right=bd, bottom=bd)
        highlight.fill = PatternFill(start_color='002aa198',
                                     end_color='002aa198',
                                     fill_type='solid')
        highlight.font = Font(color="fdf6e3")
        wb.add_named_style(highlight)

        pivot = NamedStyle(name="pivot")
        bd = Side(style='thin', color="586e75")
        pivot.border = Border(left=bd, top=bd, right=bd, bottom=bd)
        pivot.fill = PatternFill(start_color='00859900',
                                 end_color='00859900',
                                 fill_type='solid')
        pivot.font = Font(color="fdf6e3")
        wb.add_named_style(pivot)

        for i in range(1, len(problem.output_tables) + 1):
            row = 1 if i == 1 else (5 + problem.number_of_constraints) * (i - 1) + 2 + (1 if i > 2 else 0)

            ws.merge_cells(start_row=row, start_column=1, end_row=row,
                           end_column=(3 + len(problem.output_tables[i - 1].variable_names)))
            ws.cell(row=row, column=1, value=("Iter" + str(problem.output_tables[i - 1].iteration))).style = base
            row += 1

            ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=2)
            ws.cell(row=row, column=1, value="Cj").style = base
            for j in range(3, 3 + len(problem.output_tables[i - i].cj)):
                ws.cell(row=row, column=j, value=problem.output_tables[i - 1].cj[j - 3]).style = base

            ws.merge_cells(start_row=row, start_column=(3 + len(problem.output_tables[i - 1].a[0])), end_row=(row + 1),
                           end_column=(3 + len(problem.output_tables[i - 1].a[0])))
            ws.cell(row=row, column=(3 + len(problem.output_tables[i - 1].a[0])), value="b").style = base

            row += 1
            ws.merge_cells(start_row=row, start_column=1, end_row=(row + 1), end_column=1)
            ws.cell(row=row, column=1, value="Cb").style = base

            ws.cell(row=row, column=2, value="basic variables").style = base
            for j in range(3, 3 + len(problem.output_tables[i - 1].variable_names)):
                ws.cell(row=row, column=j, value=problem.output_tables[i - 1].variable_names[j - 3]).style = base

            row += 1
            ws.cell(row=row, column=2, value="zj").style = base
            for j in range(3, 3 + len(problem.output_tables[i - 1].zj)):
                ws.cell(row=row, column=j, value=problem.output_tables[i - 1].zj[j - 3]).style = base

            row += 1
            for j in range(len(problem.a)):
                ws.cell(row=(row + j), column=1, value=problem.output_tables[i - 1].cb[j]).style = base2
                ws.cell(row=(row + j), column=2, value=problem.output_tables[i - 1].basic_vars[j]).style = base2

                for k in range(len(problem.output_tables[i - 1].a[j])):
                    if j == problem.output_tables[i - 1].leaving_variable and k == problem.output_tables[
                                i - 1].entering_variable:
                        ws.cell(row=(row + j), column=(k + 3),
                                value=problem.output_tables[i - 1].a[j][k]).style = pivot
                    elif j == problem.output_tables[i - 1].leaving_variable or k == problem.output_tables[
                                i - 1].entering_variable:
                        ws.cell(row=(row + j), column=(k + 3),
                                value=problem.output_tables[i - 1].a[j][k]).style = highlight
                    else:
                        ws.cell(row=(row + j), column=(k + 3),
                                value=problem.output_tables[i - 1].a[j][k]).style = base2

                ws.cell(row=(row + j), column=(2 + len(problem.output_tables[i - 1].a[j]) + 1),
                        value=problem.output_tables[i - 1].b[j]).style = base2
            row += problem.number_of_constraints

            ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=2)
            ws.cell(row=row, column=1, value="Cj - Zj").style = base

            for j in range(3, 3 + len(problem.output_tables[i - 1].cj_minus_zj)):
                ws.cell(row=row, column=j, value=problem.output_tables[i - 1].cj_minus_zj[j - 3]).style = base
        wb.save(f)


class Simplex:
    def __init__(self, number_of_variables=None, number_of_constraints=None, cj=None, a=None, b=None,
                 constraint_types=None, variable_names=None, problem_type=None):
        self.solved = False
        self.number_of_variables = number_of_variables
        self.number_of_constraints = number_of_constraints
        self.cj = cj
        self.a = a
        self.b = b
        self.constraint_types = constraint_types
        self.variable_names = variable_names
        self.problem_type = problem_type
        self.solution = None
        self.output_tables = None

    def solve(self):
        variables = self.number_of_variables
        constraints = self.number_of_constraints
        cj = self.cj[:]
        a = [x[:] for x in self.a]
        b = self.b[:]
        constraint_types = self.constraint_types[:]
        variable_names = self.variable_names[:]
        problem_type = self.problem_type
        # Let's create the initial optimal solution
        iteration = 0
        for i in range(constraints):
            if constraint_types[i] == ">=":
                _type = 0
            elif constraint_types[i] == "<=":
                _type = 2
            else:  # =
                _type = 1
            constraint_types[i] = [constraint_types[i], _type]
            b[i] = [b[i], _type]

        for i in range(constraints):
            a[i].append(constraint_types[i][1])

        constraint_types.sort(key=lambda c: c[1])
        a.sort(key=lambda ai: ai[-1])
        b.sort(key=lambda bj: bj[-1])

        for i in range(constraints):
            a[i] = a[i][:-1]
            constraint_types[i] = constraint_types[i][0]
            b[i] = b[i][0]

        # Now I want to see the identity matrix
        gt_constraints = constraint_types.count(">=")
        variable_names += ["sr" + str(i) for i in range(1, gt_constraints + 1)]
        variable_names += ["a" + str(i) for i in range(1, gt_constraints + 1)]
        eq_constraints = constraint_types.count("=")
        variable_names += ["a" + str(i) for i in range(gt_constraints + 1, eq_constraints + gt_constraints + 1)]
        lt_constraints = constraint_types.count("<=")
        variable_names += ["sl" + str(i) for i in range(1, lt_constraints + 1)]

        cj_minus_zj = ["0.0" for x in range(variables + gt_constraints * 2 + eq_constraints + lt_constraints)]
        zj = cj_minus_zj[:]

        cj += ["0.0" for x in range(gt_constraints)]
        if problem_type == "Maximize":
            cj += ["-m" for x in range(gt_constraints + eq_constraints)]
        else:  # Minimize
            cj += ["m" for x in range(gt_constraints + eq_constraints)]
        cj += ["0.0" for x in range(lt_constraints)]

        cb = [cj[i] for i in
              range(variables + gt_constraints, variables + gt_constraints * 2 + eq_constraints + lt_constraints)]
        for i in range(len(a)):
            a[i] += [0.0 for i in range(gt_constraints * 2 + eq_constraints + lt_constraints)]

        for i in range(gt_constraints):
            for j in range(variables, variables + gt_constraints):
                if (i + variables) == j:
                    a[i][j] = -1.0

        for i in range(gt_constraints * 2 + eq_constraints + lt_constraints):
            for j in range(variables + gt_constraints,
                           variables + gt_constraints * 2 + eq_constraints + lt_constraints):
                if (i + variables + gt_constraints) == j:
                    a[i][j] = 1.0

        basic_vars = []
        for i in range(len(a)):
            for j in range(variables, len(a[i])):
                if a[i][j] == 1.0:
                    basic_vars.append(variable_names[j])

        # Now it's time to iterate until the job is done
        end = False
        special = False
        self.output_tables = []
        self.solution = []
        while not end:
            # reset z to zeroes
            zj = ["0.0" for x in range(variables + gt_constraints * 2 + eq_constraints + lt_constraints + 1)]
            # Calculate zj
            for i in range(len(a[0])):
                for j in range(len(a)):
                    new_zj = self.parse_m(zj[i] + "+" + str(a[j][i]) + str("*") + cb[j])
                    zj[i] = new_zj["rhs"] + "+" + new_zj["m_coefficient"] + "m"

            # updating the value of the solution
            for i in range(len(b)):
                aux = self.parse_m(zj[-1] + "+" + cb[i] + "*" + str(b[i]))
                zj[-1] = aux["rhs"] + "+" + aux["m_coefficient"] + "m"

            # calculate cj - zj
            for i in range(len(cj_minus_zj)):
                new_cj_minus_zj = self.parse_m(cj[i] + "-(" + zj[i] + ")")
                cj_minus_zj[i] = new_cj_minus_zj["rhs"] + ("+" if "-" not in new_cj_minus_zj["m_coefficient"] else "") + \
                                 new_cj_minus_zj["m_coefficient"] + "m"

            # check if we got to the end or we must stop (the problem is a special case)
            # Checking if we reached the stopping condition
            end = True
            if problem_type == "Maximize":
                for x in cj_minus_zj:
                    if self.parse_m(x + "> 0.0"):
                        end = False
                        break
            else:
                for x in cj_minus_zj:
                    if self.parse_m(x + "< 0.0"):
                        end = False
                        break
            # Checking if this problem is a special case
            if end:
                # 1. No feasible region
                for i in range(len(basic_vars)):
                    if "a" in basic_vars[i] and b[i] > 0.0:
                        special = True
                        self.solution = None
                        break
                # 3. Alternative solutions
                non_basic_vars = []  # Not the actual names, just the index in variable_names
                for i in range(len(variable_names)):
                    if variable_names[i] not in basic_vars:
                        non_basic_vars.append(i)

                for x in non_basic_vars:
                    if zj[x] == "0.0":
                        special = True
                        break

                self.output_tables.append(
                    SimplexOutputTable(iteration, cj, variable_names, zj, cb, basic_vars, b, cj_minus_zj,
                                       None, None, a, 0, size_hint=(None, None), width=Window.width))
                break

            # Find the entering variable
            entering_variable = cj_minus_zj[0]
            entering_variable_index = 0  # index of the entering variable
            operator = "<" if problem_type == "Maximize" else ">"
            for i in range(len(cj_minus_zj)):
                if self.parse_m(entering_variable + operator + cj_minus_zj[i]):
                    entering_variable = cj_minus_zj[i]
                    entering_variable_index = i

            # Find the leaving variable
            min_ratio = math.inf
            leaving_variable_index = None
            for i in range(len(b)):
                if a[i][entering_variable_index] > 0:
                    if min_ratio > b[i] / a[i][entering_variable_index]:
                        min_ratio = b[i] / a[i][entering_variable_index]
                        leaving_variable_index = i
            # 2. Unbounded solution
            if leaving_variable_index is None:
                special = True
                print("unbounded solution")

            # Get the pivot
            pivot = a[leaving_variable_index][entering_variable_index]

            # Get the new output table
            self.output_tables.append(
                SimplexOutputTable(iteration, cj, variable_names, zj, cb, basic_vars, b, cj_minus_zj,
                                   entering_variable_index,
                                   leaving_variable_index, a, 0, size_hint=(None, None), width=Window.width))

            # Then we pivot
            for i in range(len(a)):
                # update the bi's
                if i != leaving_variable_index:
                    b[i] -= (a[i][entering_variable_index] * b[leaving_variable_index]) / pivot

            b[leaving_variable_index] /= pivot
            # the row and column of the pivot are left untouched
            for i in range(len(a)):
                for j in range(len(a[0])):
                    if i != leaving_variable_index and j != entering_variable_index:
                        a[i][j] -= (a[i][entering_variable_index] * a[leaving_variable_index][j]) / pivot

            # Update the column of the pivot
            for i in range(len(a)):
                if i != leaving_variable_index:
                    a[i][entering_variable_index] = 0

            # Dividing the pivot's row by the pivot
            a[leaving_variable_index] = [x / pivot for x in a[leaving_variable_index]]

            # Update basic variables list and cb
            basic_vars[leaving_variable_index] = variable_names[entering_variable_index]
            cb[leaving_variable_index] = cj[entering_variable_index]

            iteration += 1
        if self.solution is not None:
            self.solution.append(float(zj[-1][:zj[-1].find("+")]))
            for x in variable_names[:variables]:
                if x in basic_vars:
                    self.solution.append(b[basic_vars.index(x)])
                else:
                    self.solution.append(0.0)
            for i in range(len(self.output_tables)):
                self.output_tables[i].number_of_iterations = len(self.output_tables)
            return self.output_tables, self.solution, basic_vars

    def add_constraint(self, coefficients, constraint_type, rhs):
        self.solved = False
        self.b.append(rhs)
        self.a.append(coefficients)
        self.constraint_types.append(constraint_type)
        self.number_of_constraints += 1
        if self.b[-1] < 0.0:
            self.b[-1] = -self.b[-1]
            self.a[-1] = [-x for x in self.a[-1]]
            if self.constraint_types[-1] == ">=":
                self.constraint_types[-1] = "<="
            elif self.constraint_types[-1] == "<=":
                self.constraint_types[-1] = ">="
        return self

    @classmethod
    def parse_m(cls, m_expression):
        _input = InputStream(m_expression)
        lexer = mLexer(_input)
        stream = CommonTokenStream(lexer)
        parser = mParser(stream)
        tree = parser.start_rule()
        visitor = mVisitor()
        return visitor.visit(tree)

    @staticmethod
    def from_simplex(s):
        new_simplex = Simplex()
        new_simplex.solved = s.solved
        new_simplex.number_of_variables = s.number_of_variables
        new_simplex.number_of_constraints = s.number_of_constraints
        new_simplex.cj = s.cj[:]
        new_simplex.a = [x[:] for x in s.a]
        new_simplex.b = s.b[:]
        new_simplex.constraint_types = s.constraint_types[:]
        new_simplex.variable_names = s.variable_names[:]
        new_simplex.problem_type = s.problem_type
        return new_simplex


class SimplexApp(App):
    def build(self):
        Window.clearcolor = [0.0, 0.16862745098039217, 0.21176470588235294, 1.0]
        gui = Gui(cols=2,
                  size_hint=(None, None), width=Window.width)
        scroll_view = ScrollView(size_hint=(None, None), size=(Window.width, Window.height), do_scroll_x=True)
        Window.bind(width=scroll_view.setter('width'))
        Window.bind(height=scroll_view.setter('height'))
        scroll_view.add_widget(gui)
        return scroll_view


if __name__ == '__main__':
    SimplexApp().run()
